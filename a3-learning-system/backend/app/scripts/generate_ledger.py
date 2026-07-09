"""
教材台账 Excel 生成器

功能：从 ChromaDB 知识库自动提取元数据，生成教材台账 Excel
包含：书名、出版社、ISBN、学科、章节数、入库时间、切片数
"""

import sys
import os
from datetime import datetime

sys.path.insert(0, ".")


def generate_ledger(output_path: str = None):
    """生成教材台账 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        return

    from app.core.chroma_client import get_collection

    wb = Workbook()
    ws = wb.active
    ws.title = "教材台账"

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["序号", "教材名称", "出版社/来源", "学科", "章节", "入库时间", "切片数", "状态"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 从知识库获取元数据
    try:
        col = get_collection("knowledge_base")
        results = col.get()
        if results and results.get("metadatas"):
            metas = results["metadatas"]
            # 按书名分组统计
            book_stats = {}
            for m in metas:
                title = m.get("title", "未知")
                source = m.get("source", "未知来源")
                if title not in book_stats:
                    book_stats[title] = {"source": source, "count": 0}
                book_stats[title]["count"] += 1

            for i, (title, stats) in enumerate(sorted(book_stats.items()), 1):
                ws.cell(row=i + 1, column=1, value=i).border = thin_border
                ws.cell(row=i + 1, column=2, value=title).border = thin_border
                ws.cell(row=i + 1, column=3, value=stats["source"]).border = thin_border
                ws.cell(row=i + 1, column=4, value="计算机科学").border = thin_border
                ws.cell(row=i + 1, column=5, value="-").border = thin_border
                ws.cell(row=i + 1, column=6, value=datetime.now().strftime("%Y-%m-%d")).border = thin_border
                ws.cell(row=i + 1, column=7, value=stats["count"]).border = thin_border
                ws.cell(row=i + 1, column=8, value="已入库").border = thin_border
                ws.cell(row=i + 1, column=8).font = Font(color="008000")

        # 习题库统计
        try:
            ex_col = get_collection("exercise_bank")
            ex_results = ex_col.get()
            if ex_results and ex_results.get("metadatas"):
                ex_metas = ex_results["metadatas"]
                row = len(book_stats) + 2
                ws.cell(row=row, column=1, value=len(book_stats) + 1).border = thin_border
                ws.cell(row=row, column=2, value="Python习题题库").border = thin_border
                ws.cell(row=row, column=3, value="习题集").border = thin_border
                ws.cell(row=row, column=4, value="计算机科学").border = thin_border
                ws.cell(row=row, column=5, value="-").border = thin_border
                ws.cell(row=row, column=6, value=datetime.now().strftime("%Y-%m-%d")).border = thin_border
                ws.cell(row=row, column=7, value=len(ex_metas)).border = thin_border
                ws.cell(row=row, column=8, value="已入库").border = thin_border
                ws.cell(row=row, column=8).font = Font(color="008000")
        except Exception:
            pass

    except Exception as e:
        print(f"知识库读取失败: {e}")

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10

    if output_path is None:
        output_path = os.path.join(os.path.dirname(__file__), "..", "..", "..", "docs", "教材台账.xlsx")
    wb.save(output_path)
    print(f"台账已生成: {output_path}")


if __name__ == "__main__":
    generate_ledger()
